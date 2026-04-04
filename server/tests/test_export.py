from openpyxl import load_workbook
from io import BytesIO

from app.services.exporter import export_ics, export_schedule


SHIFT_TYPES = [
    {"id": 1, "name": "Mañana", "start_time": "07:00", "end_time": "14:30", "effective_hours": 7.5, "priority_order": 1, "color": "#4A90D9"},
    {"id": 2, "name": "Tarde", "start_time": "14:30", "end_time": "22:00", "effective_hours": 7.5, "priority_order": 2, "color": "#E8A838"},
]

EMPLOYEES = [
    {"id": 1, "name": "Ana"},
    {"id": 2, "name": "Carlos"},
]


def test_export_creates_two_sheets():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
        {"date": "2026-03-02", "employee_id": 2, "shift_type_id": 2},
    ]
    output = export_schedule(2026, 3, EMPLOYEES, SHIFT_TYPES, assignments)

    wb = load_workbook(output)
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == "Horario"
    assert wb.sheetnames[1] == "Resumen"


def test_export_schedule_sheet_has_correct_structure():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
    ]
    output = export_schedule(2026, 3, EMPLOYEES, SHIFT_TYPES, assignments)

    wb = load_workbook(output)
    ws = wb["Horario"]

    assert ws.cell(row=1, column=1).value == "Empleado"
    assert ws.cell(row=2, column=1).value == "Ana"
    assert ws.cell(row=3, column=1).value == "Carlos"

    # March has 31 days, so column 32 should have day 31
    assert "31" in str(ws.cell(row=1, column=32).value)


def test_export_schedule_shows_shift_times():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
    ]
    output = export_schedule(2026, 3, EMPLOYEES, SHIFT_TYPES, assignments)

    wb = load_workbook(output)
    ws = wb["Horario"]

    # Day 2 is column 3 (col 1 = name, col 2 = day 1)
    cell = ws.cell(row=2, column=3)
    assert cell.value == "07:00-14:30"


def test_export_schedule_shows_free_days():
    output = export_schedule(2026, 3, EMPLOYEES, SHIFT_TYPES, [])

    wb = load_workbook(output)
    ws = wb["Horario"]

    cell = ws.cell(row=2, column=2)
    assert cell.value == "L"


def test_export_summary_has_metrics():
    assignments = [
        {"date": f"2026-03-{d:02d}", "employee_id": 1, "shift_type_id": 1}
        for d in range(2, 7)
    ]
    output = export_schedule(2026, 3, EMPLOYEES, SHIFT_TYPES, assignments)

    wb = load_workbook(output)
    ws = wb["Resumen"]

    assert ws.cell(row=1, column=1).value == "Empleado"
    assert ws.cell(row=2, column=1).value == "Ana"
    assert ws.cell(row=2, column=2).value == 5
    assert ws.cell(row=2, column=3).value == 37.5


def test_export_endpoint(client):
    client.post("/api/schedules/2026/3")
    client.put("/api/schedules/2026/3/assignments", json={
        "assignments": [
            {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
        ]
    })

    response = client.get("/api/schedules/2026/3/export")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]

    wb = load_workbook(BytesIO(response.content))
    assert len(wb.sheetnames) == 2


def test_export_endpoint_not_found(client):
    response = client.get("/api/schedules/2026/3/export")
    assert response.status_code == 404


def test_export_ics_produces_valid_vcalendar():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
        {"date": "2026-03-03", "employee_id": 1, "shift_type_id": 2},
    ]
    result = export_ics(EMPLOYEES[0], SHIFT_TYPES, assignments)

    assert result.startswith("BEGIN:VCALENDAR\r\n")
    assert result.endswith("END:VCALENDAR\r\n")
    assert "VERSION:2.0" in result
    assert result.count("BEGIN:VEVENT") == 2


def test_export_ics_event_has_correct_times_and_uid():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
    ]
    result = export_ics(EMPLOYEES[0], SHIFT_TYPES, assignments)

    assert "DTSTART:20260302T070000" in result
    assert "DTEND:20260302T143000" in result
    assert "SUMMARY:Mañana" in result
    assert "UID:20260302-1-1@shiftmaker" in result


def test_export_ics_skips_free_days():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": None},
        {"date": "2026-03-03", "employee_id": 1, "shift_type_id": 1},
    ]
    result = export_ics(EMPLOYEES[0], SHIFT_TYPES, assignments)

    assert result.count("BEGIN:VEVENT") == 1


def test_export_ics_filters_to_single_employee():
    assignments = [
        {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
        {"date": "2026-03-02", "employee_id": 2, "shift_type_id": 2},
    ]
    result = export_ics(EMPLOYEES[0], SHIFT_TYPES, assignments)

    assert result.count("BEGIN:VEVENT") == 1
    assert "SUMMARY:Mañana" in result


def test_export_ics_endpoint(client):
    client.post("/api/schedules/2026/3")
    client.put("/api/schedules/2026/3/assignments", json={
        "assignments": [
            {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
        ]
    })

    response = client.get("/api/schedules/2026/3/export-ics/1")
    assert response.status_code == 200
    assert "text/calendar" in response.headers["content-type"]

    content = response.text
    assert "BEGIN:VCALENDAR" in content
    assert "BEGIN:VEVENT" in content


def test_export_ics_endpoint_filename(client):
    client.post("/api/schedules/2026/3")
    client.put("/api/schedules/2026/3/assignments", json={
        "assignments": [
            {"date": "2026-03-02", "employee_id": 1, "shift_type_id": 1},
        ]
    })

    response = client.get("/api/schedules/2026/3/export-ics/1")
    disposition = response.headers["content-disposition"]

    assert "Ana" in disposition
    assert "2026-03" in disposition
    assert ".ics" in disposition


def test_export_ics_endpoint_schedule_not_found(client):
    response = client.get("/api/schedules/2026/3/export-ics/1")
    assert response.status_code == 404


def test_export_ics_endpoint_employee_not_found(client):
    client.post("/api/schedules/2026/3")

    response = client.get("/api/schedules/2026/3/export-ics/999")
    assert response.status_code == 404
