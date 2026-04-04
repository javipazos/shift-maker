import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def export_schedule(
    year: int,
    month: int,
    employees: list[dict],
    shift_types: list[dict],
    assignments: list[dict],
) -> BytesIO:
    wb = Workbook()
    _build_schedule_sheet(wb.active, year, month, employees, shift_types, assignments)
    _build_summary_sheet(wb.create_sheet("Resumen"), employees, shift_types, assignments)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_schedule_sheet(
    ws, year, month, employees, shift_types, assignments
):
    ws.title = "Horario"
    days_in_month = calendar.monthrange(year, month)[1]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    weekend_fill = PatternFill(start_color="F0F0F5", end_color="F0F0F5", fill_type="solid")
    header_font = Font(bold=True, size=9)
    cell_font = Font(size=8)
    center = Alignment(horizontal="center", vertical="center")

    shift_map = {st["id"]: st for st in shift_types}

    # Header row: employee name + day numbers
    ws.cell(row=1, column=1, value="Empleado").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border
    ws.column_dimensions["A"].width = 18

    day_names = ["L", "M", "X", "J", "V", "S", "D"]

    for d in range(1, days_in_month + 1):
        col = d + 1
        date_str = f"{year}-{month:02d}-{d:02d}"
        dow = calendar.weekday(year, month, d)
        cell = ws.cell(row=1, column=col, value=f"{day_names[dow]}\n{d}")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.fill = weekend_fill if dow >= 5 else header_fill
        ws.column_dimensions[get_column_letter(col)].width = 7

    # Employee rows
    for row_idx, emp in enumerate(employees, start=2):
        ws.cell(row=row_idx, column=1, value=emp["name"]).font = Font(bold=True, size=9)
        ws.cell(row=row_idx, column=1).border = thin_border

        for d in range(1, days_in_month + 1):
            col = d + 1
            date_str = f"{year}-{month:02d}-{d:02d}"
            dow = calendar.weekday(year, month, d)

            assignment = _find_assignment(assignments, date_str, emp["id"])
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = center
            cell.font = cell_font

            if dow >= 5:
                cell.fill = weekend_fill

            if assignment and assignment["shift_type_id"] is not None:
                shift = shift_map.get(assignment["shift_type_id"])
                if shift:
                    cell.value = f"{shift['start_time']}-{shift['end_time']}"
                    color = shift["color"].lstrip("#")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(size=8, color="FFFFFF", bold=True)
            else:
                cell.value = "L"
                cell.font = Font(size=8, color="999999")

    # Coverage row
    cov_row = len(employees) + 2
    ws.cell(row=cov_row, column=1, value="Cobertura").font = Font(bold=True, italic=True, size=9)
    ws.cell(row=cov_row, column=1).border = thin_border

    for d in range(1, days_in_month + 1):
        col = d + 1
        date_str = f"{year}-{month:02d}-{d:02d}"
        count = sum(
            1 for a in assignments
            if a["date"] == date_str and a["shift_type_id"] is not None
        )
        cell = ws.cell(row=cov_row, column=col, value=count)
        cell.font = Font(bold=True, size=9)
        cell.alignment = center
        cell.border = thin_border


def _build_summary_sheet(ws, employees, shift_types, assignments):
    ws.title = "Resumen"

    headers = ["Empleado", "Días trabajados", "Horas totales", "Horas/semana (media)"]
    header_font = Font(bold=True, size=10)

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 20

    shift_map = {st["id"]: st for st in shift_types}

    for row_idx, emp in enumerate(employees, start=2):
        emp_assignments = [
            a for a in assignments
            if a["employee_id"] == emp["id"] and a["shift_type_id"] is not None
        ]
        total_hours = sum(
            shift_map.get(a["shift_type_id"], {}).get("effective_hours", 0)
            for a in emp_assignments
        )
        days_worked = len(emp_assignments)
        # Rough weekly average (month ≈ 4.3 weeks)
        weekly_avg = total_hours / 4.3 if total_hours > 0 else 0

        ws.cell(row=row_idx, column=1, value=emp["name"])
        ws.cell(row=row_idx, column=2, value=days_worked)
        ws.cell(row=row_idx, column=3, value=round(total_hours, 1))
        ws.cell(row=row_idx, column=4, value=round(weekly_avg, 1))


def export_ics(
    employee: dict,
    shift_types: list[dict],
    assignments: list[dict],
) -> str:
    shift_map = {st["id"]: st for st in shift_types}
    emp_shifts = [
        a for a in assignments
        if a["employee_id"] == employee["id"] and a["shift_type_id"] is not None
    ]

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Shift Maker//ES",
        "CALSCALE:GREGORIAN",
    ]

    for assignment in emp_shifts:
        shift = shift_map.get(assignment["shift_type_id"])
        if not shift:
            continue
        date_str = assignment["date"].replace("-", "")
        start = shift["start_time"].replace(":", "")
        end = shift["end_time"].replace(":", "")
        lines += [
            "BEGIN:VEVENT",
            f"DTSTART:{date_str}T{start}00",
            f"DTEND:{date_str}T{end}00",
            f"SUMMARY:{shift['name']}",
            "END:VEVENT",
        ]

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def _find_assignment(assignments, date, emp_id):
    for a in assignments:
        if a["date"] == date and a["employee_id"] == emp_id:
            return a
    return None
