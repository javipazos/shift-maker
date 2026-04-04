import re
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook


@dataclass
class ImportResult:
    assignments: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


class ImportError(Exception):
    pass


def parse_schedule_xlsx(
    file: BytesIO,
    employees: list[dict],
    shift_types: list[dict],
    year: int,
    month: int,
) -> ImportResult:
    result = ImportResult()

    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result.errors.append("El archivo está vacío")
        return result

    header_row = rows[0]
    day_columns = _parse_header_columns(header_row)

    if not day_columns:
        result.errors.append("No se encontraron columnas con números de día en la cabecera")
        return result

    emp_lookup = _build_employee_lookup(employees)
    shift_lookup = _build_shift_lookup(shift_types)

    for row in rows[1:]:
        if not row or not row[0]:
            continue

        emp_name = str(row[0]).strip()
        employee = emp_lookup.get(emp_name.lower())

        if not employee:
            result.warnings.append(f"Empleado no encontrado: '{emp_name}'")
            continue

        for col_idx, day_num in day_columns.items():
            date_str = f"{year}-{month:02d}-{day_num:02d}"
            cell_value = row[col_idx] if col_idx < len(row) else None
            shift_type_id = _resolve_shift(cell_value, shift_lookup)

            if shift_type_id == -1:
                result.warnings.append(
                    f"Turno desconocido '{cell_value}' para {emp_name} el día {day_num}"
                )
                shift_type_id = None

            result.assignments.append({
                "date": date_str,
                "employee_id": employee["id"],
                "shift_type_id": shift_type_id,
            })

    return result


def _parse_header_columns(header: tuple) -> dict[int, int]:
    """Map column index -> day number from the header row."""
    columns = {}
    for i, cell in enumerate(header):
        if cell is None or i == 0:
            continue
        day = _extract_day_number(cell)
        if day is not None:
            columns[i] = day
    return columns


def _extract_day_number(value) -> int | None:
    """Extract day number from header cell (e.g. 1, '1', 'L\\n1')."""
    if isinstance(value, (int, float)):
        n = int(value)
        if 1 <= n <= 31:
            return n
        return None

    text = str(value).strip()
    # Handle "L\n1" format from export
    match = re.search(r"\d+", text)
    if match:
        n = int(match.group())
        if 1 <= n <= 31:
            return n
    return None


def _build_employee_lookup(employees: list[dict]) -> dict[str, dict]:
    return {e["name"].lower(): e for e in employees}


def _build_shift_lookup(shift_types: list[dict]) -> dict[str, int]:
    """Build lookup: lowercase name / time range -> shift_type_id."""
    lookup: dict[str, int] = {}
    for st in shift_types:
        lookup[st["name"].lower()] = st["id"]
        time_key = f"{st['start_time']}-{st['end_time']}"
        lookup[time_key] = st["id"]
    return lookup


def _resolve_shift(value, shift_lookup: dict[str, int]) -> int | None:
    """Return shift_type_id, None for free day, or -1 for unknown."""
    if value is None:
        return None

    text = str(value).strip()
    if not text or text.upper() == "L":
        return None

    key = text.lower()
    if key in shift_lookup:
        return shift_lookup[key]

    # -1 signals unknown shift
    return -1
